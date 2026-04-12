from __future__ import annotations

import json
import os
import pickle
import re
from pathlib import Path
from typing import Any

import baseline_recipe_selection_lib

try:
    from openpyxl import load_workbook
except Exception as exc:  # pragma: no cover - exercised only when dependency is missing.
    load_workbook = None  # type: ignore[assignment]
    _OPENPYXL_IMPORT_ERROR: Exception | None = exc
else:
    _OPENPYXL_IMPORT_ERROR = None

MASTER_MAP_FILENAME = "material_label_map_w_chemicalname_分類_性質數據.xlsx"
MASTER_MAP_SHEET = "Sheet1"
DEFAULT_RATIO_MIN = 1.5
DEFAULT_RATIO_MAX = 2.5
DEFAULT_HISTORY_LOWER = 10
DEFAULT_HISTORY_UPPER = 90
DEFAULT_EXAMPLE_LIMIT = 5
VISCOSITY_FIELDS = [
    "25℃粘度（mPa.s）",
    "40℃粘度（mPa.s）",
    "50℃粘度（mPa.s）",
    "70℃粘度（mPa.s）",
    "100℃粘度（mPa.s）",
    "150℃粘度（mPa.s）",
    "200℃粘度（mPa.s）",
]


def dependency_error() -> str | None:
    if _OPENPYXL_IMPORT_ERROR is None:
        return None
    return f"prepreg_b_ratio_guard requires openpyxl: {_OPENPYXL_IMPORT_ERROR}"


def clean_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def normalize_material_name(value: Any) -> str:
    return clean_text(value).upper()


def as_float(value: Any) -> float | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = clean_text(value)
    if not text:
        return None
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return None
    try:
        return float(match.group(0))
    except Exception:
        return None


def clamp_int(value: Any, default: int, *, lower: int, upper: int) -> int:
    parsed = int(as_float(value) or default)
    return max(lower, min(upper, parsed))


def plugin_root() -> Path:
    root = os.environ.get("CLAWD_PLUGIN_ROOT")
    if root:
        return Path(root).expanduser()
    return Path(__file__).resolve().parents[1]


def workspace_root() -> Path:
    env_root = clean_text(os.environ.get("OPENCLAW_WORKSPACE"))
    if env_root:
        return Path(env_root).expanduser()
    return Path(__file__).resolve().parents[6]


def default_material_data_dir() -> Path:
    return plugin_root().parent / "materials-search" / "data" / "material"


def resolve_master_map_path(payload: dict[str, Any]) -> Path:
    workbooks = payload.get("workbooks") if isinstance(payload.get("workbooks"), dict) else {}
    workbook_path = clean_text(workbooks.get("master_map_path")) if isinstance(workbooks, dict) else ""
    if workbook_path:
        return Path(workbook_path).expanduser()
    material_data_dir = clean_text(payload.get("material_data_dir"))
    if material_data_dir:
        return Path(material_data_dir).expanduser() / MASTER_MAP_FILENAME
    return default_material_data_dir() / MASTER_MAP_FILENAME


def default_history_path() -> Path:
    candidates: list[Path] = []
    env_path = clean_text(os.environ.get("PREPREG_B_RATIO_HISTORY_PATH"))
    if env_path:
        candidates.append(Path(env_path).expanduser())
    workspace = workspace_root()
    fact_cache = workspace / "knowledge" / "fact-cache"
    candidates.extend(
        [
            *sorted(fact_cache.glob("recipe-store-v3-*.pkl")),
            *sorted(fact_cache.glob("recipe-store-v2-*.pkl")),
            *sorted(fact_cache.glob("recipe-store-*.pkl")),
        ]
    )
    candidates.append(baseline_recipe_selection_lib.resolve_default_export_path())
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return candidates[0] if candidates else baseline_recipe_selection_lib.resolve_default_export_path()


def resolve_history_path(payload: dict[str, Any]) -> Path:
    explicit = clean_text(payload.get("history_path") or payload.get("export_path"))
    if explicit:
        return Path(explicit).expanduser()
    return default_history_path()


def source_file_info(path: Path, *, kind: str, extra: dict[str, Any] | None = None) -> dict[str, Any]:
    info: dict[str, Any] = {"path": str(path), "kind": kind, "exists": path.exists()}
    if path.exists():
        stat = path.stat()
        info.update({"size_bytes": int(stat.st_size), "mtime_ns": int(stat.st_mtime_ns)})
    if extra:
        info.update(extra)
    return info


def base_response(action: str) -> dict[str, Any]:
    return {
        "status": "ok",
        "action": action,
        "source_files": [],
        "query_summary": {"action": action},
        "material_resolution": [],
        "role_summary": {},
        "input_summary": {},
        "hard_filter": {},
        "history_summary": {},
        "recommendation_summary": {},
        "model_summary": {},
        "example_records": [],
        "suggestions": [],
        "warnings": [],
        "errors": [],
    }


def apply_status(response: dict[str, Any]) -> dict[str, Any]:
    if response["errors"]:
        response["status"] = "error"
    elif response["warnings"]:
        response["status"] = "warning"
    else:
        response["status"] = "ok"
    return response


def require_openpyxl() -> None:
    if load_workbook is None:
        raise RuntimeError(dependency_error() or "openpyxl is unavailable")


def load_master_store(path: Path) -> dict[str, Any]:
    require_openpyxl()
    if not path.exists():
        raise FileNotFoundError(f"Material workbook not found: {path}")
    workbook = load_workbook(path, data_only=True, read_only=True)  # type: ignore[misc]
    if MASTER_MAP_SHEET not in workbook.sheetnames:
        raise ValueError(f"Workbook {path} missing required sheet: {MASTER_MAP_SHEET}")
    sheet = workbook[MASTER_MAP_SHEET]
    rows = sheet.iter_rows(values_only=True)
    try:
        header = [clean_text(cell) for cell in next(rows)]
    except StopIteration:
        header = []
    entries: list[dict[str, Any]] = []
    index: dict[str, int] = {}
    for row_number, row in enumerate(rows, start=2):
        raw = {header[idx]: row[idx] for idx in range(min(len(header), len(row))) if header[idx]}
        material = clean_text(raw.get("材料編號"))
        if not material:
            continue
        entry = {
            "row_number": row_number,
            "material": material,
            "normalized_material": normalize_material_name(material),
            "category": clean_text(raw.get("分類")),
            "family": clean_text(raw.get("固化劑體系")),
            "appearance": clean_text(raw.get("外观")),
            "solid_content": raw.get("固体含量（%）"),
            "melting_point": raw.get("熔点（℃）"),
            "softening_point": raw.get("软化点（℃）"),
            "viscosities": {field: raw.get(field) for field in VISCOSITY_FIELDS if raw.get(field) not in (None, "")},
            "row": raw,
        }
        entries.append(entry)
        index.setdefault(entry["normalized_material"], len(entries) - 1)
    return {
        "entries": entries,
        "index": index,
        "source_file": source_file_info(path, kind="material_master", extra={"sheet": MASTER_MAP_SHEET}),
    }


def _contains_any(text: str, terms: list[str]) -> bool:
    lowered = text.lower()
    return any(term in text or term in lowered for term in terms)


def classify_material_state(entry: dict[str, Any] | None) -> dict[str, Any]:
    if not entry:
        return {
            "state": "unknown",
            "confidence": "low",
            "evidence": ["material_not_found_in_master_map"],
            "category": None,
            "family": None,
            "appearance": None,
        }

    appearance = clean_text(entry.get("appearance"))
    category = clean_text(entry.get("category"))
    family = clean_text(entry.get("family"))
    melting_point = as_float(entry.get("melting_point"))
    softening_point = as_float(entry.get("softening_point"))
    solid_content = as_float(entry.get("solid_content"))
    viscosities = entry.get("viscosities") if isinstance(entry.get("viscosities"), dict) else {}
    evidence: list[str] = []

    if _contains_any(appearance, ["粉", "颗粒", "粒", "片", "晶", "powder", "flake", "granule", "crystal"]):
        evidence.append(f"appearance={appearance or 'solid_marker'}")
        return {
            "state": "solid",
            "confidence": "high",
            "evidence": evidence,
            "category": category or None,
            "family": family or None,
            "appearance": appearance or None,
        }
    if _contains_any(appearance, ["液", "油", "浆", "liquid", "viscous"]):
        evidence.append(f"appearance={appearance or 'liquid_marker'}")
        return {
            "state": "liquid",
            "confidence": "high",
            "evidence": evidence,
            "category": category or None,
            "family": family or None,
            "appearance": appearance or None,
        }
    if any(token in category for token in ["稀释剂", "偶联剂", "消泡剂", "色浆", "色膏"]):
        evidence.append(f"category={category}")
        return {
            "state": "liquid",
            "confidence": "high",
            "evidence": evidence,
            "category": category or None,
            "family": family or None,
            "appearance": appearance or None,
        }
    if melting_point is not None and melting_point >= 40:
        evidence.append(f"melting_point={melting_point}")
        return {
            "state": "solid",
            "confidence": "high",
            "evidence": evidence,
            "category": category or None,
            "family": family or None,
            "appearance": appearance or None,
        }
    if softening_point is not None and softening_point >= 40:
        evidence.append(f"softening_point={softening_point}")
        return {
            "state": "solid",
            "confidence": "high",
            "evidence": evidence,
            "category": category or None,
            "family": family or None,
            "appearance": appearance or None,
        }
    if solid_content is not None and solid_content >= 95:
        evidence.append(f"solid_content={solid_content}")
        return {
            "state": "solid",
            "confidence": "medium",
            "evidence": evidence,
            "category": category or None,
            "family": family or None,
            "appearance": appearance or None,
        }
    if viscosities:
        evidence.append("viscosity_fields_present")
        return {
            "state": "liquid",
            "confidence": "medium",
            "evidence": evidence,
            "category": category or None,
            "family": family or None,
            "appearance": appearance or None,
        }
    if "树脂" in category:
        evidence.append(f"category={category}")
        return {
            "state": "liquid",
            "confidence": "low",
            "evidence": evidence,
            "category": category or None,
            "family": family or None,
            "appearance": appearance or None,
        }
    if any(token in category for token in ["填料", "阻燃剂", "固化剂", "促进剂"]):
        evidence.append(f"category={category}")
        return {
            "state": "solid",
            "confidence": "low",
            "evidence": evidence,
            "category": category or None,
            "family": family or None,
            "appearance": appearance or None,
        }
    return {
        "state": "unknown",
        "confidence": "low",
        "evidence": ["insufficient_physical_state_evidence"],
        "category": category or None,
        "family": family or None,
        "appearance": appearance or None,
    }


def infer_component_role(entry: dict[str, Any] | None, state_info: dict[str, Any]) -> dict[str, Any]:
    category = clean_text(state_info.get("category") or (entry.get("category") if entry else ""))
    family = clean_text(state_info.get("family") or (entry.get("family") if entry else ""))
    state = clean_text(state_info.get("state"))
    evidence: list[str] = []

    def build(role_key: str, role_label: str, confidence: str, reason: str) -> dict[str, Any]:
        return {
            "role_key": role_key,
            "role_label": role_label,
            "role_confidence": confidence,
            "role_reason": reason,
        }

    if state == "solid" and "促进剂/固化剂" in category:
        evidence.append(f"category={category}")
        return build("latent_curing_package_solid", "潜伏固化包固体", "high", ",".join(evidence))
    if state == "solid" and "固化剂" in category:
        evidence.append(f"category={category}")
        if "预浸料" in family:
            evidence.append(f"family={family}")
        return build("curing_agent_solid", "固化剂固体", "high", ",".join(evidence))
    if state == "solid" and "促进剂" in category:
        evidence.append(f"category={category}")
        return build("accelerator_solid", "促进剂固体", "high", ",".join(evidence))
    if state == "solid" and any(token in category for token in ["填料", "阻燃剂"]):
        evidence.append(f"category={category}")
        return build("functional_filler_solid", "功能粉体", "medium", ",".join(evidence))
    if state == "liquid" and "树脂" in category:
        evidence.append(f"category={category}")
        return build("carrier_resin_liquid", "载体树脂液体", "high", ",".join(evidence))
    if state == "liquid" and any(token in category for token in ["稀释剂", "偶联剂", "消泡剂", "色浆", "色膏"]):
        evidence.append(f"category={category}")
        return build("process_additive_liquid", "工艺助剂液体", "high", ",".join(evidence))
    if state == "liquid":
        if category:
            evidence.append(f"category={category}")
        return build("liquid_component", "液体组分", "medium", ",".join(evidence) or "state=liquid")
    if state == "solid":
        if category:
            evidence.append(f"category={category}")
        return build("solid_component", "固体组分", "medium", ",".join(evidence) or "state=solid")
    return build("unknown_component", "待确认组分", "low", "state=unknown")


def normalize_components(payload: dict[str, Any]) -> tuple[list[dict[str, Any]], list[str]]:
    warnings: list[str] = []
    raw_components = payload.get("components")
    if not isinstance(raw_components, list):
        return [], ["components must be an array."]
    normalized: list[dict[str, Any]] = []
    seen: set[str] = set()
    for index, item in enumerate(raw_components):
        if not isinstance(item, dict):
            warnings.append(f"components[{index}] is not an object and was ignored.")
            continue
        material = clean_text(item.get("material"))
        if not material:
            warnings.append(f"components[{index}] missing material and was ignored.")
            continue
        amount = as_float(item.get("amount"))
        proportion = as_float(item.get("proportion"))
        if amount is None and proportion is None:
            warnings.append(f"components[{index}] for {material} requires amount or proportion.")
            continue
        key = normalize_material_name(material)
        if key in seen:
            warnings.append(f"Duplicate component {material} ignored after first occurrence.")
            continue
        seen.add(key)
        normalized.append(
            {
                "material": material,
                "normalized_material": key,
                "amount": amount,
                "proportion": proportion,
            }
        )
    return normalized, warnings


def resolve_total_b_amount(payload: dict[str, Any], components: list[dict[str, Any]]) -> tuple[float | None, list[str]]:
    warnings: list[str] = []
    explicit_total = as_float(payload.get("b_total_amount"))
    if explicit_total is None:
        explicit_total = as_float(payload.get("mix_ratio_hardener"))

    amount_sum = sum(component["amount"] for component in components if component.get("amount") is not None)
    has_amounts = any(component.get("amount") is not None for component in components)
    has_proportions = any(component.get("proportion") is not None for component in components)

    if has_amounts and not has_proportions:
        if explicit_total is None:
            return amount_sum, warnings
        if abs(amount_sum - explicit_total) > max(0.2, explicit_total * 0.05):
            warnings.append(
                f"b_total_amount={explicit_total} differs from component amount sum={round(amount_sum, 6)}; using b_total_amount."
            )
        return explicit_total, warnings

    if has_proportions and not has_amounts:
        if explicit_total is None:
            return None, ["b_total_amount is required when components are provided as proportions."]
        proportion_sum = sum(component["proportion"] for component in components if component.get("proportion") is not None)
        if abs(proportion_sum - 100.0) > 1.0:
            warnings.append(f"component proportion sum is {round(proportion_sum, 4)}, not 100.")
        return explicit_total, warnings

    if has_amounts and has_proportions:
        warnings.append("components should use either amount or proportion; amount values were preferred where present.")
        if explicit_total is None:
            return amount_sum if amount_sum > 0 else None, warnings
        return explicit_total, warnings

    return explicit_total, warnings


def enrich_input_components(
    components: list[dict[str, Any]],
    *,
    total_b_amount: float,
    master_store: dict[str, Any],
) -> list[dict[str, Any]]:
    enriched: list[dict[str, Any]] = []
    uses_proportion = any(component.get("proportion") is not None for component in components)
    for component in components:
        amount = component.get("amount")
        if amount is None and uses_proportion:
            amount = total_b_amount * float(component.get("proportion") or 0.0) / 100.0
        entry = master_store["entries"][master_store["index"][component["normalized_material"]]] if component["normalized_material"] in master_store["index"] else None
        state_info = classify_material_state(entry)
        role_info = infer_component_role(entry, state_info)
        enriched.append(
            {
                "material": component["material"],
                "normalized_material": component["normalized_material"],
                "amount": round(float(amount or 0.0), 6),
                "proportion": component.get("proportion"),
                "state": state_info["state"],
                "state_confidence": state_info["confidence"],
                "state_evidence": list(state_info["evidence"]),
                "category": state_info["category"],
                "family": state_info["family"],
                "appearance": state_info["appearance"],
                "role_key": role_info["role_key"],
                "role_label": role_info["role_label"],
                "role_confidence": role_info["role_confidence"],
                "role_reason": role_info["role_reason"],
            }
        )
    return enriched


def summarize_roles(components: list[dict[str, Any]]) -> dict[str, Any]:
    counts: dict[str, int] = {}
    amounts: dict[str, float] = {}
    labels: dict[str, str] = {}
    for component in components:
        role_key = clean_text(component.get("role_key"))
        if not role_key:
            continue
        counts[role_key] = counts.get(role_key, 0) + 1
        amounts[role_key] = amounts.get(role_key, 0.0) + float(component.get("amount") or 0.0)
        labels[role_key] = clean_text(component.get("role_label")) or role_key
    ordered_roles = sorted(
        (
            {
                "role_key": role_key,
                "role_label": labels.get(role_key) or role_key,
                "component_count": counts[role_key],
                "total_amount": round(amounts[role_key], 6),
            }
            for role_key in counts
        ),
        key=lambda item: (-float(item["total_amount"]), item["role_key"]),
    )
    return {
        "role_count": len(ordered_roles),
        "roles": ordered_roles,
    }


def summarize_components(components: list[dict[str, Any]], total_b_amount: float | None) -> dict[str, Any]:
    solid_amount = sum(component["amount"] for component in components if component.get("state") == "solid")
    liquid_amount = sum(component["amount"] for component in components if component.get("state") == "liquid")
    unknown_amount = sum(component["amount"] for component in components if component.get("state") == "unknown")
    total_amount = float(total_b_amount if total_b_amount is not None else (solid_amount + liquid_amount + unknown_amount))
    ratio = None if solid_amount <= 0 else liquid_amount / solid_amount
    return {
        "component_count": len(components),
        "total_b_amount": round(total_amount, 6),
        "solid_amount": round(solid_amount, 6),
        "liquid_amount": round(liquid_amount, 6),
        "unknown_amount": round(unknown_amount, 6),
        "solid_share": round((solid_amount / total_amount), 6) if total_amount > 0 else None,
        "liquid_share": round((liquid_amount / total_amount), 6) if total_amount > 0 else None,
        "unknown_share": round((unknown_amount / total_amount), 6) if total_amount > 0 else None,
        "liquid_to_solid_ratio": round(ratio, 6) if ratio is not None else None,
    }


def ratio_guidance(total_b_amount: float, ratio_min: float, ratio_max: float) -> dict[str, float]:
    solid_min = total_b_amount / (1.0 + ratio_max)
    solid_max = total_b_amount / (1.0 + ratio_min)
    liquid_min = total_b_amount - solid_max
    liquid_max = total_b_amount - solid_min
    return {
        "solid_min": round(solid_min, 6),
        "solid_max": round(solid_max, 6),
        "liquid_min": round(liquid_min, 6),
        "liquid_max": round(liquid_max, 6),
    }


def build_hard_filter(summary: dict[str, Any], *, ratio_min: float, ratio_max: float) -> tuple[dict[str, Any], list[str]]:
    suggestions: list[str] = []
    total_b_amount = float(summary.get("total_b_amount") or 0.0)
    solid_amount = float(summary.get("solid_amount") or 0.0)
    liquid_amount = float(summary.get("liquid_amount") or 0.0)
    unknown_amount = float(summary.get("unknown_amount") or 0.0)
    ratio = summary.get("liquid_to_solid_ratio")
    filter_result = {
        "rule": {
            "liquid_to_solid_min": ratio_min,
            "liquid_to_solid_max": ratio_max,
            "description": "预浸料 B 剂液体:固体建议保持在 1.5-2.5:1；液体不能少于固体，否则人工难以搅拌。",
        },
        "allowed_amount_window": ratio_guidance(total_b_amount, ratio_min, ratio_max) if total_b_amount > 0 else {},
        "actual_ratio": ratio,
        "status": "blocked",
        "message": "",
    }
    if unknown_amount > 0:
        filter_result["status"] = "blocked"
        filter_result["message"] = "存在未识别相态的物料，无法做硬筛选。"
        suggestions.append("先补齐 unknown 物料的主数据相态，再执行硬筛选。")
        return filter_result, suggestions
    if solid_amount <= 0:
        filter_result["status"] = "blocked"
        filter_result["message"] = "固体用量为 0，无法计算液体:固体比例。"
        suggestions.append("至少提供一个可识别为 solid 的 B 剂组分。")
        return filter_result, suggestions
    if ratio is None:
        filter_result["status"] = "blocked"
        filter_result["message"] = "缺少液体/固体数据，无法计算比例。"
        return filter_result, suggestions
    if ratio_min <= float(ratio) <= ratio_max:
        filter_result["status"] = "pass"
        filter_result["message"] = "液体:固体比例落在硬筛选范围内。"
        return filter_result, suggestions
    if float(ratio) < ratio_min:
        required_liquid = max(0.0, ratio_min * solid_amount - liquid_amount)
        max_solid = liquid_amount / ratio_min if ratio_min > 0 else solid_amount
        suggestions.append(f"当前液体偏少。若固体不变，至少再增加 {round(required_liquid, 6)} 份液体。")
        suggestions.append(f"若总液体不变，固体应降到不高于 {round(max_solid, 6)} 份。")
        filter_result["status"] = "fail"
        filter_result["message"] = "液体:固体比例低于下限，B 剂偏干，存在难搅拌风险。"
        return filter_result, suggestions
    required_solid = max(0.0, liquid_amount / ratio_max - solid_amount)
    max_liquid = ratio_max * solid_amount
    suggestions.append(f"当前液体偏多。若液体不变，至少再增加 {round(required_solid, 6)} 份固体。")
    suggestions.append(f"若固体不变，液体应降到不高于 {round(max_liquid, 6)} 份。")
    filter_result["status"] = "fail"
    filter_result["message"] = "液体:固体比例高于上限，B 剂偏稀。"
    return filter_result, suggestions


def load_history_records(path: Path) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    if not path.exists():
        raise FileNotFoundError(f"History file not found: {path}")
    if path.suffix.casefold() == ".pkl":
        with path.open("rb") as handle:
            parsed = pickle.load(handle)
        if not isinstance(parsed, dict):
            raise ValueError(f"Unsupported pickle payload: {path}")
        records = parsed.get("records")
        if not isinstance(records, list):
            raise ValueError(f"Recipe cache missing records list: {path}")
        return records, {
            "source_kind": "pickle_cache",
            "dataset_count": len(records),
            "cache_signature": parsed.get("cache_signature"),
            "source_files": parsed.get("source_files"),
        }
    records = baseline_recipe_selection_lib.load_baseline_raw_records_from_export(path)
    return records, {"source_kind": "baseline_export", "dataset_count": len(records)}


def extract_prepreg_history_rows(records: list[dict[str, Any]], master_store: dict[str, Any]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for record in records:
        if "payload" in record:
            payload = baseline_recipe_selection_lib.get_payload(record)
            experiment_id = clean_text(payload.get("experimentId") or record.get("id"))
        else:
            payload = record
            experiment_id = clean_text(record.get("experimentId") or record.get("experiment_id") or record.get("point_id"))
        family = clean_text(payload.get("hardener_system_family"))
        if family != "预浸料体系":
            continue
        material_groups = payload.get("materialGroups") or payload.get("materialGroup") or {}
        if not isinstance(material_groups, dict):
            continue
        hardener_group = material_groups.get("hardener")
        if not isinstance(hardener_group, dict):
            continue
        components = hardener_group.get("components")
        if not isinstance(components, dict) or not components:
            continue
        total_b_amount = as_float(hardener_group.get("proportion"))
        if total_b_amount is None:
            mix_ratio_parsed = payload.get("mix_ratio_parsed") if isinstance(payload.get("mix_ratio_parsed"), dict) else {}
            total_b_amount = as_float(mix_ratio_parsed.get("hardener"))
        if total_b_amount is None or total_b_amount <= 0:
            total_b_amount = as_float(payload.get("mix_ratio_hardener"))
        if total_b_amount is None or total_b_amount <= 0:
            continue
        component_rows: list[dict[str, Any]] = []
        solid_amount = 0.0
        liquid_amount = 0.0
        unknown_amount = 0.0
        for material, value in components.items():
            amount_share = as_float(value)
            if amount_share is None:
                continue
            amount = total_b_amount * amount_share / 100.0
            normalized = normalize_material_name(material)
            entry = master_store["entries"][master_store["index"][normalized]] if normalized in master_store["index"] else None
            state_info = classify_material_state(entry)
            component_rows.append(
                {
                    "material": clean_text(material),
                    "state": state_info["state"],
                    "amount_share": round(amount_share, 6),
                    "amount": round(amount, 6),
                }
            )
            if state_info["state"] == "solid":
                solid_amount += amount
            elif state_info["state"] == "liquid":
                liquid_amount += amount
            else:
                unknown_amount += amount
        if solid_amount <= 0 or liquid_amount <= 0:
            continue
        ratio = liquid_amount / solid_amount
        rows.append(
            {
                "experiment_id": experiment_id or None,
                "total_b_amount": round(total_b_amount, 6),
                "solid_amount": round(solid_amount, 6),
                "liquid_amount": round(liquid_amount, 6),
                "unknown_amount": round(unknown_amount, 6),
                "liquid_to_solid_ratio": round(ratio, 6),
                "component_materials": sorted(clean_text(material) for material in components.keys() if clean_text(material)),
                "component_rows": component_rows,
            }
        )
    return rows


def percentile(values: list[float], point: float) -> float | None:
    if not values:
        return None
    ordered = sorted(values)
    if len(ordered) == 1:
        return ordered[0]
    position = max(0.0, min(100.0, float(point))) / 100.0 * (len(ordered) - 1)
    lower_index = int(position)
    upper_index = min(len(ordered) - 1, lower_index + 1)
    fraction = position - lower_index
    return ordered[lower_index] + (ordered[upper_index] - ordered[lower_index]) * fraction


def build_history_summary(
    rows: list[dict[str, Any]],
    *,
    lower_percentile: int,
    upper_percentile: int,
    total_b_amount: float | None,
) -> dict[str, Any]:
    solid_values = [float(row["solid_amount"]) for row in rows]
    liquid_values = [float(row["liquid_amount"]) for row in rows]
    ratio_values = [float(row["liquid_to_solid_ratio"]) for row in rows]
    summary = {
        "dataset_count": len(rows),
        "total_b_amount": {
            "p10": round(percentile([float(row["total_b_amount"]) for row in rows], 10) or 0.0, 6) if rows else None,
            "p50": round(percentile([float(row["total_b_amount"]) for row in rows], 50) or 0.0, 6) if rows else None,
            "p90": round(percentile([float(row["total_b_amount"]) for row in rows], 90) or 0.0, 6) if rows else None,
        },
        "solid_amount_window": {
            "lower_percentile": lower_percentile,
            "upper_percentile": upper_percentile,
            "min": round(min(solid_values), 6) if solid_values else None,
            "max": round(max(solid_values), 6) if solid_values else None,
            "recommended_min": round(percentile(solid_values, lower_percentile) or 0.0, 6) if solid_values else None,
            "recommended_median": round(percentile(solid_values, 50) or 0.0, 6) if solid_values else None,
            "recommended_max": round(percentile(solid_values, upper_percentile) or 0.0, 6) if solid_values else None,
        },
        "liquid_amount_window": {
            "recommended_min": round(percentile(liquid_values, lower_percentile) or 0.0, 6) if liquid_values else None,
            "recommended_median": round(percentile(liquid_values, 50) or 0.0, 6) if liquid_values else None,
            "recommended_max": round(percentile(liquid_values, upper_percentile) or 0.0, 6) if liquid_values else None,
        },
        "ratio_distribution": {
            "min": round(min(ratio_values), 6) if ratio_values else None,
            "p10": round(percentile(ratio_values, 10) or 0.0, 6) if ratio_values else None,
            "p50": round(percentile(ratio_values, 50) or 0.0, 6) if ratio_values else None,
            "p90": round(percentile(ratio_values, 90) or 0.0, 6) if ratio_values else None,
            "max": round(max(ratio_values), 6) if ratio_values else None,
        },
    }
    if total_b_amount is not None and total_b_amount > 0:
        summary["hard_rule_from_total_b"] = ratio_guidance(total_b_amount, DEFAULT_RATIO_MIN, DEFAULT_RATIO_MAX)
        solid_window = summary["solid_amount_window"]
        hist_min = solid_window.get("recommended_min")
        hist_max = solid_window.get("recommended_max")
        hard_rule = summary["hard_rule_from_total_b"]
        if hist_min is not None and hist_max is not None:
            combined_min = max(float(hist_min), float(hard_rule["solid_min"]))
            combined_max = min(float(hist_max), float(hard_rule["solid_max"]))
            summary["combined_solid_guidance"] = {
                "history_min": hist_min,
                "history_max": hist_max,
                "hard_rule_min": hard_rule["solid_min"],
                "hard_rule_max": hard_rule["solid_max"],
                "intersection_min": round(combined_min, 6),
                "intersection_max": round(combined_max, 6),
                "has_overlap": combined_min <= combined_max,
            }
    return summary


def intersect_windows(left_min: float | None, left_max: float | None, right_min: float | None, right_max: float | None) -> dict[str, Any]:
    if None in {left_min, left_max, right_min, right_max}:
        return {
            "history_min": left_min,
            "history_max": left_max,
            "hard_rule_min": right_min,
            "hard_rule_max": right_max,
            "intersection_min": None,
            "intersection_max": None,
            "has_overlap": False,
        }
    low = max(float(left_min), float(right_min))
    high = min(float(left_max), float(right_max))
    return {
        "history_min": round(float(left_min), 6),
        "history_max": round(float(left_max), 6),
        "hard_rule_min": round(float(right_min), 6),
        "hard_rule_max": round(float(right_max), 6),
        "intersection_min": round(low, 6),
        "intersection_max": round(high, 6),
        "has_overlap": low <= high,
    }


def build_recommendation_summary(
    *,
    input_summary: dict[str, Any],
    hard_filter: dict[str, Any],
    history_summary: dict[str, Any],
) -> dict[str, Any]:
    hard_rule_window = hard_filter.get("allowed_amount_window") if isinstance(hard_filter.get("allowed_amount_window"), dict) else {}
    solid_history = history_summary.get("solid_amount_window") if isinstance(history_summary.get("solid_amount_window"), dict) else {}
    liquid_history = history_summary.get("liquid_amount_window") if isinstance(history_summary.get("liquid_amount_window"), dict) else {}

    solid_range = intersect_windows(
        solid_history.get("recommended_min"),
        solid_history.get("recommended_max"),
        hard_rule_window.get("solid_min"),
        hard_rule_window.get("solid_max"),
    )
    liquid_range = intersect_windows(
        liquid_history.get("recommended_min"),
        liquid_history.get("recommended_max"),
        hard_rule_window.get("liquid_min"),
        hard_rule_window.get("liquid_max"),
    )

    actual_solid = as_float(input_summary.get("solid_amount"))
    actual_liquid = as_float(input_summary.get("liquid_amount"))
    hard_status = clean_text(hard_filter.get("status"))
    has_overlap = bool(solid_range.get("has_overlap")) and bool(liquid_range.get("has_overlap"))
    within_recommended_window = (
        has_overlap
        and actual_solid is not None
        and actual_liquid is not None
        and float(solid_range["intersection_min"]) <= actual_solid <= float(solid_range["intersection_max"])
        and float(liquid_range["intersection_min"]) <= actual_liquid <= float(liquid_range["intersection_max"])
    )

    operable = hard_status == "pass" and has_overlap
    if hard_status == "blocked":
        operability_status = "blocked"
        message = "当前输入无法完成可操作性判断。"
    elif hard_status == "fail":
        operability_status = "not_operable"
        message = "当前配方未通过液固硬筛选，不建议直接操作。"
    elif not has_overlap:
        operability_status = "needs_review"
        message = "历史窗口与硬规则窗口没有形成稳定交集，建议人工复核。"
    elif within_recommended_window:
        operability_status = "operable"
        message = "当前配方落在历史+硬规则共同窗口内，可操作。"
    else:
        operability_status = "operable_with_adjustment"
        message = "当前配方通过硬筛选，可操作，但未落在历史推荐绝对量窗口内，建议继续调优。"

    return {
        "operable": operable,
        "operability_status": operability_status,
        "message": message,
        "recommended_solid_amount_range": {
            "min": solid_range.get("intersection_min"),
            "max": solid_range.get("intersection_max"),
        },
        "recommended_liquid_amount_range": {
            "min": liquid_range.get("intersection_min"),
            "max": liquid_range.get("intersection_max"),
        },
        "recommended_ratio_range": {
            "min": hard_filter.get("rule", {}).get("liquid_to_solid_min") if isinstance(hard_filter.get("rule"), dict) else None,
            "max": hard_filter.get("rule", {}).get("liquid_to_solid_max") if isinstance(hard_filter.get("rule"), dict) else None,
            "actual": input_summary.get("liquid_to_solid_ratio"),
        },
        "window_breakdown": {
            "solid": solid_range,
            "liquid": liquid_range,
        },
    }


def build_example_records(
    history_rows: list[dict[str, Any]],
    *,
    current_components: list[dict[str, Any]],
    current_summary: dict[str, Any],
    limit: int,
) -> list[dict[str, Any]]:
    if not history_rows or limit <= 0:
        return []
    requested_materials = {component["material"] for component in current_components}
    requested_ratio = as_float(current_summary.get("liquid_to_solid_ratio"))
    requested_solid = as_float(current_summary.get("solid_amount"))
    scored: list[tuple[tuple[float, float, float, str], dict[str, Any]]] = []
    for row in history_rows:
        candidate_materials = set(row.get("component_materials") or [])
        overlap = len(candidate_materials & requested_materials)
        ratio_gap = abs(float(row["liquid_to_solid_ratio"]) - float(requested_ratio or 0.0))
        solid_gap = abs(float(row["solid_amount"]) - float(requested_solid or 0.0))
        scored.append(
            (
                (-float(overlap), ratio_gap, solid_gap, clean_text(row.get("experiment_id"))),
                {
                    "experiment_id": row.get("experiment_id"),
                    "overlap_material_count": overlap,
                    "component_materials": row.get("component_materials"),
                    "total_b_amount": row.get("total_b_amount"),
                    "solid_amount": row.get("solid_amount"),
                    "liquid_amount": row.get("liquid_amount"),
                    "liquid_to_solid_ratio": row.get("liquid_to_solid_ratio"),
                },
            )
        )
    scored.sort(key=lambda item: item[0])
    return [item for _, item in scored[:limit]]


def schema_payload() -> dict[str, Any]:
    return {
        "supported_actions": ["schema", "analyze"],
        "history_scope": "预浸料体系 B 剂历史记录",
        "ratio_rule": {
            "liquid_to_solid_min": DEFAULT_RATIO_MIN,
            "liquid_to_solid_max": DEFAULT_RATIO_MAX,
        },
        "component_input_modes": [
            {"mode": "amount", "require_b_total_amount": False},
            {"mode": "proportion", "require_b_total_amount": True},
        ],
        "notes": [
            "该工具只针对预浸料 B 剂设计。",
            "历史统计从本地 recipe cache pickle 或 baseline export JSON 读取。",
            "物料液体/固体相态从 material_label_map_w_chemicalname_分類_性質數據.xlsx 推断。",
            "输出会包含组分角色识别、推荐绝对量区间，以及适合模型消费的简化 JSON 摘要。",
        ],
    }


def build_model_summary(
    *,
    material_resolution: list[dict[str, Any]],
    input_summary: dict[str, Any],
    recommendation_summary: dict[str, Any],
    hard_filter: dict[str, Any],
    suggestions: list[str],
) -> dict[str, Any]:
    operability_status = clean_text(recommendation_summary.get("operability_status"))
    hard_filter_status = clean_text(hard_filter.get("status"))
    next_action_key = "manual_review"
    next_action_message = clean_text(recommendation_summary.get("message"))
    if operability_status == "blocked":
        next_action_key = "resolve_unknowns"
    elif operability_status == "not_operable":
        if hard_filter_status == "fail":
            actual_ratio = as_float(input_summary.get("liquid_to_solid_ratio"))
            ratio_rule = hard_filter.get("rule") if isinstance(hard_filter.get("rule"), dict) else {}
            ratio_min = as_float(ratio_rule.get("liquid_to_solid_min"))
            ratio_max = as_float(ratio_rule.get("liquid_to_solid_max"))
            if actual_ratio is not None and ratio_min is not None and actual_ratio < ratio_min:
                next_action_key = "add_liquid_or_reduce_solid"
            elif actual_ratio is not None and ratio_max is not None and actual_ratio > ratio_max:
                next_action_key = "add_solid_or_reduce_liquid"
            else:
                next_action_key = "rebalance_formula"
        else:
            next_action_key = "rebalance_formula"
    elif operability_status == "needs_review":
        next_action_key = "manual_review"
    elif operability_status == "operable_with_adjustment":
        next_action_key = "tune_to_window"
    elif operability_status == "operable":
        next_action_key = "use_as_is"
    if suggestions:
        next_action_message = suggestions[0]

    return {
        "decision": {
            "operable": bool(recommendation_summary.get("operable")),
            "status": recommendation_summary.get("operability_status"),
            "reason": recommendation_summary.get("message"),
            "hard_filter_status": hard_filter.get("status"),
        },
        "target_window": {
            "solid_amount_min": recommendation_summary.get("recommended_solid_amount_range", {}).get("min")
            if isinstance(recommendation_summary.get("recommended_solid_amount_range"), dict) else None,
            "solid_amount_max": recommendation_summary.get("recommended_solid_amount_range", {}).get("max")
            if isinstance(recommendation_summary.get("recommended_solid_amount_range"), dict) else None,
            "liquid_amount_min": recommendation_summary.get("recommended_liquid_amount_range", {}).get("min")
            if isinstance(recommendation_summary.get("recommended_liquid_amount_range"), dict) else None,
            "liquid_amount_max": recommendation_summary.get("recommended_liquid_amount_range", {}).get("max")
            if isinstance(recommendation_summary.get("recommended_liquid_amount_range"), dict) else None,
            "ratio_min": recommendation_summary.get("recommended_ratio_range", {}).get("min")
            if isinstance(recommendation_summary.get("recommended_ratio_range"), dict) else None,
            "ratio_max": recommendation_summary.get("recommended_ratio_range", {}).get("max")
            if isinstance(recommendation_summary.get("recommended_ratio_range"), dict) else None,
        },
        "current": {
            "total_b_amount": input_summary.get("total_b_amount"),
            "solid_amount": input_summary.get("solid_amount"),
            "liquid_amount": input_summary.get("liquid_amount"),
            "unknown_amount": input_summary.get("unknown_amount"),
            "liquid_to_solid_ratio": input_summary.get("liquid_to_solid_ratio"),
        },
        "components": [
            {
                "material": component.get("material"),
                "amount": component.get("amount"),
                "state": component.get("state"),
                "role": component.get("role_key"),
                "confidence": {
                    "state": component.get("state_confidence"),
                    "role": component.get("role_confidence"),
                },
            }
            for component in material_resolution
        ],
        "next_action": {
            "action": next_action_key,
            "message": next_action_message,
        },
    }


def run_tool(payload: dict[str, Any]) -> dict[str, Any]:
    action = clean_text(payload.get("action") or "schema").lower()
    response = base_response(action)
    response["query_summary"].update(
        {
            "history_path": str(resolve_history_path(payload)),
            "master_map_path": str(resolve_master_map_path(payload)),
            "hardener_system_type": clean_text(payload.get("hardener_system_type") or "PREPREG").upper(),
        }
    )

    if action == "schema":
        response["query_summary"]["schema"] = schema_payload()
        return response

    if action != "analyze":
        response["errors"].append(f"Unsupported action: {action}")
        return apply_status(response)

    if clean_text(payload.get("hardener_system_type") or "PREPREG").upper() not in {"", "PREPREG"}:
        response["errors"].append("prepreg_b_ratio_guard currently supports hardener_system_type=PREPREG only.")
        return apply_status(response)

    ratio_limits = payload.get("ratio_limits") if isinstance(payload.get("ratio_limits"), dict) else {}
    ratio_min = as_float(ratio_limits.get("liquid_to_solid_min")) or DEFAULT_RATIO_MIN
    ratio_max = as_float(ratio_limits.get("liquid_to_solid_max")) or DEFAULT_RATIO_MAX
    if ratio_min <= 0 or ratio_max <= 0 or ratio_min >= ratio_max:
        response["errors"].append("ratio_limits must satisfy 0 < liquid_to_solid_min < liquid_to_solid_max.")
        return apply_status(response)

    lower_percentile = clamp_int(payload.get("history_lower_percentile"), DEFAULT_HISTORY_LOWER, lower=0, upper=49)
    upper_percentile = clamp_int(payload.get("history_upper_percentile"), DEFAULT_HISTORY_UPPER, lower=51, upper=100)
    example_limit = clamp_int(payload.get("example_limit"), DEFAULT_EXAMPLE_LIMIT, lower=0, upper=20)

    try:
        master_store = load_master_store(resolve_master_map_path(payload))
    except Exception as exc:
        response["errors"].append(f"{type(exc).__name__}: {exc}")
        return apply_status(response)

    components, component_warnings = normalize_components(payload)
    response["warnings"].extend(component_warnings)
    if not components:
        response["errors"].append("No usable components were provided.")
        return apply_status(response)

    total_b_amount, total_warnings = resolve_total_b_amount(payload, components)
    response["warnings"].extend(total_warnings)
    if total_b_amount is None or total_b_amount <= 0:
        response["errors"].append("Unable to resolve a positive b_total_amount.")
        return apply_status(response)

    resolved_components = enrich_input_components(components, total_b_amount=total_b_amount, master_store=master_store)
    input_summary = summarize_components(resolved_components, total_b_amount)
    hard_filter, filter_suggestions = build_hard_filter(input_summary, ratio_min=ratio_min, ratio_max=ratio_max)

    history_path = resolve_history_path(payload)
    response["source_files"] = [
        master_store["source_file"],
        source_file_info(history_path, kind="history_source"),
    ]

    try:
        raw_history_records, history_meta = load_history_records(history_path)
        history_rows = extract_prepreg_history_rows(raw_history_records, master_store)
    except Exception as exc:
        response["errors"].append(f"{type(exc).__name__}: {exc}")
        return apply_status(response)

    history_summary = build_history_summary(
        history_rows,
        lower_percentile=lower_percentile,
        upper_percentile=upper_percentile,
        total_b_amount=total_b_amount,
    )
    history_summary["history_source"] = history_meta

    if not history_rows:
        response["warnings"].append("No usable prepreg B-side history rows were extracted from the history source.")

    response["material_resolution"] = resolved_components
    response["role_summary"] = summarize_roles(resolved_components)
    response["input_summary"] = input_summary
    response["hard_filter"] = hard_filter
    response["history_summary"] = history_summary
    response["recommendation_summary"] = build_recommendation_summary(
        input_summary=input_summary,
        hard_filter=hard_filter,
        history_summary=history_summary,
    )
    response["example_records"] = build_example_records(
        history_rows,
        current_components=resolved_components,
        current_summary=input_summary,
        limit=example_limit,
    )
    response["suggestions"].extend(filter_suggestions)

    hard_filter_status = clean_text(hard_filter.get("status"))
    hard_filter_message = clean_text(hard_filter.get("message"))
    if hard_filter_status in {"fail", "blocked"} and hard_filter_message:
        response["warnings"].append(hard_filter_message)

    if hard_filter.get("status") == "pass" and history_rows:
        recommendation_summary = response["recommendation_summary"]
        solid_window = recommendation_summary.get("window_breakdown", {}).get("solid") if isinstance(recommendation_summary.get("window_breakdown"), dict) else {}
        if isinstance(solid_window, dict) and solid_window.get("has_overlap") is False:
            response["warnings"].append("当前总 B 量对应的硬规则固体窗口与历史固体窗口没有重叠，建议复核总量设定。")
            response["suggestions"].append(
                f"历史推荐固体范围约为 {solid_window.get('history_min')} - {solid_window.get('history_max')}，"
                f"但按当前总 B 量硬规则允许范围仅为 {solid_window.get('hard_rule_min')} - {solid_window.get('hard_rule_max')}。"
            )

    response["model_summary"] = build_model_summary(
        material_resolution=resolved_components,
        input_summary=input_summary,
        recommendation_summary=response["recommendation_summary"],
        hard_filter=hard_filter,
        suggestions=response["suggestions"],
    )

    return apply_status(response)
