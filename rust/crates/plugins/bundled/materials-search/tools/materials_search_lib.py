from __future__ import annotations

import os
import re
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except Exception as exc:  # pragma: no cover - exercised only when dependency is missing.
    load_workbook = None  # type: ignore[assignment]
    _OPENPYXL_IMPORT_ERROR: Exception | None = exc
else:
    _OPENPYXL_IMPORT_ERROR = None


def plugin_root() -> Path:
    root = os.environ.get("CLAWD_PLUGIN_ROOT")
    if root:
        return Path(root).expanduser()
    return Path(__file__).resolve().parents[1]


def default_material_data_dir() -> Path:
    return plugin_root() / "data" / "material"

DEFAULT_DATA_DIR = Path(
    os.environ.get(
        "OPENCLAW_MATERIAL_DATA_DIR",
        os.environ.get("GD_MATERIAL_DATA_DIR", str(default_material_data_dir())),
    )
).expanduser()
MASTER_MAP_FILENAME = "material_label_map_w_chemicalname_分類_性質數據.xlsx"
B_CLASSIFICATION_FILENAME = "原材料分类.xlsx"
MASTER_MAP_SHEET = "Sheet1"
B_CLASSIFICATION_SHEET = "B 剂"

TEXT_FIELDS = {
    "材料編號",
    "化學名",
    "CAS number",
    "IUPAC Name",
    "SMILES",
    "固化劑體系",
    "分類",
    "外观",
}

IDENTITY_FIELDS = {
    "材料編號",
    "化學名",
    "CAS number",
    "PubChem CID",
    "IUPAC Name",
    "SMILES",
    "固化劑體系",
    "分類",
}

FAMILY_ALIASES = {
    "胺类": "胺类体系",
    "胺類": "胺类体系",
    "胺类体系": "胺类体系",
    "胺類體系": "胺类体系",
    "酸酐": "酸酐类体系",
    "酸酐体系": "酸酐类体系",
    "酸酐类": "酸酐类体系",
    "酸酐类体系": "酸酐类体系",
    "酸酐類": "酸酐类体系",
    "酸酐類體系": "酸酐类体系",
    "预浸料": "预浸料体系",
    "預浸料": "预浸料体系",
    "预浸料体系": "预浸料体系",
    "預浸料體系": "预浸料体系",
    "双氰胺": "预浸料体系",
    "雙氰胺": "预浸料体系",
    "双氰胺体系": "预浸料体系",
    "雙氰胺體系": "预浸料体系",
    "双氰胺（预浸料）": "预浸料体系",
    "双氰胺(预浸料)": "预浸料体系",
    "预浸料/双氰胺": "预浸料体系",
}

CROSS_SYSTEM_EXCEPTION_RULES = [
    {
        "rule_id": "PREPREG_ACCELERATOR_CROSS_USE",
        "source_family": "预浸料体系",
        "allowed_target_families": ["胺类体系", "酸酐类体系"],
        "required_function_tags": ["促进剂"],
        "message": "双氰胺（预浸料）体系促进剂可部分用于胺类/酸酐配方，但只能按条件放行，不能视为默认兼容。",
    },
    {
        "rule_id": "AMINE_ACCELERATOR_OR_CURING_AGENT_CROSS_USE",
        "source_family": "胺类体系",
        "allowed_target_families": ["预浸料体系", "酸酐类体系"],
        "required_function_tags": ["促进剂", "固化剂"],
        "message": "胺类体系促进剂/固化剂可部分用于双氰胺（预浸料）/酸酐配方，但属于高风险条件放行。",
    },
    {
        "rule_id": "ANHYDRIDE_ACCELERATOR_CROSS_USE",
        "source_family": "酸酐类体系",
        "allowed_target_families": ["预浸料体系", "胺类体系"],
        "required_function_tags": ["促进剂"],
        "message": "酸酐体系促进剂可部分用于双氰胺（预浸料）/胺类配方，但只能按条件放行。",
    },
]

_DASHES = r"[\u2010\u2011\u2012\u2013\u2014\u2212\uFE58\uFE63\uFF0D]"


def dependency_error() -> str | None:
    if _OPENPYXL_IMPORT_ERROR is None:
        return None
    return f"materials_search requires openpyxl: {_OPENPYXL_IMPORT_ERROR}"


def clean_text(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    text = text.replace("，", ",").replace("／", "/")
    return re.sub(r"\s+", " ", text)


def normalize_token(value: Any) -> str:
    text = clean_text(value)
    if not text:
        return ""
    text = re.sub(_DASHES, "-", text)
    text = re.sub(r"[\s_\-:/,.'\"()（）]+", "", text)
    return text.casefold()


def normalize_material_token(value: Any) -> str:
    text = clean_text(value)
    if not text:
        return ""
    text = re.sub(_DASHES, "-", text)
    return re.sub(r"\s+", "", text).upper()


def tokenize(value: Any) -> list[str]:
    text = clean_text(value)
    if not text:
        return []
    return [token.casefold() for token in re.findall(r"[\u4e00-\u9fffA-Za-z0-9.\-_/]+", text)]


def as_string_list(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, str):
        value = [value]
    if not isinstance(value, list):
        return []
    return [clean_text(item) for item in value if clean_text(item)]


def as_bool(value: Any) -> bool:
    return bool(value) if isinstance(value, bool) else False


def limit_from(value: Any, default: int) -> int:
    if isinstance(value, bool):
        return default
    try:
        limit = int(value)
    except (TypeError, ValueError):
        return default
    return max(0, min(limit, 100))


def _value_as_output(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        return stripped or None
    return value


def resolve_paths(payload: dict[str, Any]) -> dict[str, Path]:
    data_dir = Path(clean_text(payload.get("data_dir")) or DEFAULT_DATA_DIR)
    workbooks = payload.get("workbooks") if isinstance(payload.get("workbooks"), dict) else {}
    master_path = clean_text(workbooks.get("master_map_path")) if isinstance(workbooks, dict) else ""
    b_path = clean_text(workbooks.get("b_classification_path")) if isinstance(workbooks, dict) else ""
    return {
        "master_map": Path(master_path) if master_path else data_dir / MASTER_MAP_FILENAME,
        "b_classification": Path(b_path) if b_path else data_dir / B_CLASSIFICATION_FILENAME,
    }


def source_file_info(path: Path, sheet: str) -> dict[str, Any]:
    info: dict[str, Any] = {
        "path": str(path),
        "sheet": sheet,
        "exists": path.exists(),
    }
    if path.exists():
        stat = path.stat()
        info.update({"size_bytes": int(stat.st_size), "mtime_ns": int(stat.st_mtime_ns)})
    return info


def base_response(action: str, paths: dict[str, Path] | None = None) -> dict[str, Any]:
    paths = paths or {}
    source_files = []
    if "master_map" in paths:
        source_files.append(source_file_info(paths["master_map"], MASTER_MAP_SHEET))
    if "b_classification" in paths:
        source_files.append(source_file_info(paths["b_classification"], B_CLASSIFICATION_SHEET))
    return {
        "status": "ok",
        "action": action,
        "source_files": source_files,
        "query_summary": {"action": action},
        "results": [],
        "safety_results": [],
        "profile": {},
        "warnings": [],
        "errors": [],
    }


def require_openpyxl() -> None:
    if load_workbook is None:
        raise RuntimeError(dependency_error() or "openpyxl is unavailable")


def load_sheet(path: Path, sheet_name: str):
    require_openpyxl()
    if not path.exists():
        raise FileNotFoundError(f"Workbook not found: {path}")
    workbook = load_workbook(path, data_only=True, read_only=True)  # type: ignore[misc]
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Workbook {path} missing required sheet: {sheet_name}")
    return workbook[sheet_name]


def load_master_store(path: Path) -> dict[str, Any]:
    sheet = load_sheet(path, MASTER_MAP_SHEET)
    rows = sheet.iter_rows(values_only=True)
    try:
        header = [clean_text(cell) for cell in next(rows)]
    except StopIteration:
        header = []
    entries: list[dict[str, Any]] = []
    material_index: dict[str, list[int]] = {}
    cas_index: dict[str, list[int]] = {}

    for row_number, row in enumerate(rows, start=2):
        raw = {header[idx]: _value_as_output(row[idx]) for idx in range(min(len(header), len(row))) if header[idx]}
        material_code = clean_text(raw.get("材料編號"))
        if not material_code:
            continue
        chemical_name = clean_text(raw.get("化學名"))
        cas_number = clean_text(raw.get("CAS number"))
        text_parts = [str(raw.get(field) or "") for field in TEXT_FIELDS]
        entry = {
            "row_number": row_number,
            "sheet": MASTER_MAP_SHEET,
            "material_code": material_code,
            "material_token": normalize_token(material_code),
            "chemical_name": chemical_name or None,
            "chemical_token": normalize_token(chemical_name),
            "cas_number": cas_number or None,
            "cas_token": normalize_token(cas_number),
            "family": _value_as_output(raw.get("固化劑體系")),
            "category": _value_as_output(raw.get("分類")),
            "appearance": _value_as_output(raw.get("外观")),
            "row": raw,
            "search_tokens": sorted(set(token for part in text_parts for token in tokenize(part))),
        }
        index = len(entries)
        entries.append(entry)
        if entry["material_token"]:
            material_index.setdefault(entry["material_token"], []).append(index)
        if entry["cas_token"]:
            cas_index.setdefault(entry["cas_token"], []).append(index)

    return {
        "source_files": [source_file_info(path, MASTER_MAP_SHEET)],
        "sheet": MASTER_MAP_SHEET,
        "columns": header,
        "dataset_count": len(entries),
        "entries": entries,
        "indexes": {"material": material_index, "cas": cas_index},
    }


def normalize_family_label(value: Any) -> str | None:
    text = clean_text(value)
    if not text:
        return None
    if text == "酸酐/胺类":
        return text
    return FAMILY_ALIASES.get(text, text if text in {"胺类体系", "酸酐类体系", "预浸料体系"} else None)


def parse_function_tags(value: Any) -> list[str]:
    text = clean_text(value)
    tags: list[str] = []
    if not text:
        return tags
    if re.search(r"促进剂|催化剂|加速剂", text, re.IGNORECASE):
        tags.append("促进剂")
    if re.search(r"固化剂", text, re.IGNORECASE):
        tags.append("固化剂")
    return tags


def build_material_aliases(raw_material: str) -> list[str]:
    aliases: list[str] = []
    normalized = normalize_material_token(raw_material)
    if normalized:
        aliases.append(normalized)
        if normalized.endswith("*"):
            aliases.append(normalized.rstrip("*"))
    for part in re.split(r"[／/]", str(raw_material or "")):
        token = normalize_material_token(part)
        if token:
            aliases.append(token)
            if token.endswith("*"):
                aliases.append(token.rstrip("*"))
    out: list[str] = []
    seen: set[str] = set()
    for alias in aliases:
        if alias not in seen:
            seen.add(alias)
            out.append(alias)
    return out


def load_b_store(path: Path) -> dict[str, Any]:
    sheet = load_sheet(path, B_CLASSIFICATION_SHEET)
    entries: list[dict[str, Any]] = []
    index: dict[str, list[int]] = {}

    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        raw_material = clean_text(row[0] if len(row) > 0 else None)
        if not raw_material:
            continue
        raw_family = clean_text(row[1] if len(row) > 1 else None)
        raw_function = clean_text(row[2] if len(row) > 2 else None)
        raw_error_note = clean_text(row[3] if len(row) > 3 else None)
        entry = {
            "sheet": B_CLASSIFICATION_SHEET,
            "row_number": row_number,
            "material": raw_material,
            "normalized_material": normalize_material_token(raw_material),
            "aliases": build_material_aliases(raw_material),
            "family_raw": raw_family or None,
            "family": normalize_family_label(raw_family),
            "function_raw": raw_function or None,
            "function_tags": parse_function_tags(raw_function),
            "classification_error_note": raw_error_note or None,
        }
        entry_index = len(entries)
        entries.append(entry)
        for alias in entry["aliases"]:
            index.setdefault(alias, []).append(entry_index)

    return {
        "source_files": [source_file_info(path, B_CLASSIFICATION_SHEET)],
        "dataset_count": len(entries),
        "entries": entries,
        "index": index,
    }


def numeric_prefix(value: Any) -> float | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    match = re.search(r"-?\d+(?:\.\d+)?", str(value))
    return float(match.group(0)) if match else None


def normalized_property(column: str, value: Any) -> dict[str, Any] | None:
    number = numeric_prefix(value)
    if number is None:
        return None
    column_lc = column.casefold()
    value_lc = str(value).casefold()
    if "密度" in column or "density" in column_lc:
        if "kg/m" in value_lc:
            return {"value": number / 1000.0, "unit": "g/cm3", "source": "kg/m3_to_g/cm3"}
        return {"value": number, "unit": "g/cm3", "source": "as_recorded"}
    if "粘度" in column or "黏度" in column or "viscosity" in column_lc:
        if "pa" in value_lc and "mpa" not in value_lc:
            return {"value": number * 1000.0, "unit": "mPa*s", "source": "Pa*s_to_mPa*s"}
        if any(unit in value_lc for unit in ["cps", "cp", "mpa"]):
            return {"value": number, "unit": "mPa*s", "source": "cP_or_mPa*s"}
    if "环氧当量" in column or "環氧當量" in column or "eew" in column_lc:
        return {"value": number, "unit": "g/eq", "source": "as_recorded"}
    if "ahew" in column_lc:
        return {"value": number, "unit": "g/eq", "source": "as_recorded"}
    if "胺值" in column:
        return {"value": number, "unit": "mgKOH/g", "source": "as_recorded"}
    return None


def entry_to_output(entry: dict[str, Any], *, include_all_fields: bool = False, score: float | None = None) -> dict[str, Any]:
    properties = {
        key: value
        for key, value in entry["row"].items()
        if key not in IDENTITY_FIELDS and value not in (None, "")
    }
    normalized_properties = {
        key: converted
        for key, value in properties.items()
        if (converted := normalized_property(key, value)) is not None
    }
    out: dict[str, Any] = {
        "source": "material_master",
        "row_number": entry["row_number"],
        "sheet": entry["sheet"],
        "material_code": entry["material_code"],
        "chemical_name": entry["chemical_name"],
        "cas_number": entry["cas_number"],
        "iupac_name": entry["row"].get("IUPAC Name"),
        "smiles": entry["row"].get("SMILES"),
        "molecular_formula": entry["row"].get("分子式"),
        "molecular_weight": entry["row"].get("分子量(Mw)(g/mol)"),
        "hardener_family": entry["family"],
        "category": entry["category"],
        "appearance": entry["appearance"],
        "property_summary": properties,
        "normalized_property_summary": normalized_properties,
    }
    if score is not None:
        out["score"] = score
    if include_all_fields:
        out["all_fields"] = dict(entry["row"])
    return out


def master_schema(store: dict[str, Any]) -> dict[str, Any]:
    family_counts: dict[str, int] = {}
    category_counts: dict[str, int] = {}
    property_columns: list[str] = []
    for column in store["columns"]:
        nonempty = sum(1 for entry in store["entries"] if entry["row"].get(column) not in (None, ""))
        if column == "固化劑體系":
            for entry in store["entries"]:
                if entry.get("family"):
                    family_counts[str(entry["family"])] = family_counts.get(str(entry["family"]), 0) + 1
        if column == "分類":
            for entry in store["entries"]:
                if entry.get("category"):
                    category_counts[str(entry["category"])] = category_counts.get(str(entry["category"]), 0) + 1
        if nonempty and column not in IDENTITY_FIELDS:
            property_columns.append(column)
    return {
        "source": "material_master",
        "dataset_count": store["dataset_count"],
        "sheet": store["sheet"],
        "key_fields": ["材料編號", "化學名", "CAS number", "IUPAC Name", "SMILES", "固化劑體系", "分類"],
        "property_columns": property_columns,
        "top_families": family_counts,
        "top_categories": category_counts,
    }


def b_schema(store: dict[str, Any]) -> dict[str, Any]:
    return {
        "source": "b_classification",
        "dataset_count": store["dataset_count"],
        "sheet": B_CLASSIFICATION_SHEET,
        "supported_target_families": ["胺类体系", "酸酐类体系", "预浸料体系"],
        "hard_rules": [
            "仅允许三类跨体系例外：预浸料促进剂 -> 胺类/酸酐；胺类促进剂或固化剂 -> 预浸料/酸酐；酸酐促进剂 -> 预浸料/胺类。",
            "除上述例外外，酸酐类材料与胺类材料不得混用。",
            "“部分可用”一律按条件放行处理，不得解释成默认兼容。",
        ],
        "cross_system_exception_rules": CROSS_SYSTEM_EXCEPTION_RULES,
    }


def lookup_master(store: dict[str, Any], materials: list[str], cas_numbers: list[str], include_all_fields: bool) -> tuple[list[dict[str, Any]], list[str]]:
    warnings: list[str] = []
    seen: set[int] = set()
    ordered: list[int] = []
    for material in materials:
        token = normalize_token(material)
        matches = store["indexes"]["material"].get(token, [])
        if not matches:
            warnings.append(f"材料主数据未命中材料编号: {material}")
        for index in matches:
            if index not in seen:
                seen.add(index)
                ordered.append(index)
    for cas in cas_numbers:
        token = normalize_token(cas)
        matches = store["indexes"]["cas"].get(token, [])
        if not matches:
            warnings.append(f"材料主数据未命中 CAS: {cas}")
        for index in matches:
            if index not in seen:
                seen.add(index)
                ordered.append(index)
    return [entry_to_output(store["entries"][index], include_all_fields=include_all_fields) for index in ordered], warnings


def search_master(store: dict[str, Any], query: str, field: str, limit: int) -> list[dict[str, Any]]:
    query_text = clean_text(query)
    query_token = normalize_token(query_text)
    query_terms = tokenize(query_text)
    if not query_text:
        return []

    scored: list[tuple[tuple[float, int, str], dict[str, Any], float]] = []
    for entry in store["entries"]:
        score = 0.0
        haystacks = []
        if field in {"all", "material"}:
            haystacks.append(("material", entry["material_code"], entry["material_token"]))
        if field in {"all", "chemical"}:
            haystacks.append(("chemical", entry["chemical_name"] or "", entry["chemical_token"]))
        if field in {"all", "cas"}:
            haystacks.append(("cas", entry["cas_number"] or "", entry["cas_token"]))
        if field in {"all", "text"}:
            joined = " ".join(str(entry["row"].get(key) or "") for key in TEXT_FIELDS)
            haystacks.append(("text", joined, normalize_token(joined)))

        matched = False
        for _, display_text, token in haystacks:
            if not display_text:
                continue
            display_norm = clean_text(display_text)
            if query_token and token == query_token:
                score += 100
                matched = True
            elif query_token and query_token in token:
                score += 50
                matched = True
            elif query_text.casefold() in display_norm.casefold():
                score += 30
                matched = True
            else:
                ratio = SequenceMatcher(None, query_token, token).ratio() if query_token and token else 0.0
                if ratio >= 0.78:
                    score += ratio * 25
                    matched = True

        overlap = len(set(query_terms).intersection(entry["search_tokens"])) if query_terms else 0
        if overlap:
            score += overlap * 10
            matched = True
        if matched:
            scored.append(((-score, len(entry["material_code"]), entry["material_code"]), entry, score))

    scored.sort(key=lambda item: item[0])
    return [entry_to_output(entry, score=round(score, 3)) for _, entry, score in scored[:limit]]


def filter_master(store: dict[str, Any], family: str | None, category: str | None, has_fields: list[str], limit: int) -> list[dict[str, Any]]:
    family_filter = clean_text(family)
    category_filter = clean_text(category)
    required_fields = [clean_text(field) for field in has_fields if clean_text(field)]
    results: list[dict[str, Any]] = []
    for entry in store["entries"]:
        if family_filter and clean_text(entry.get("family")) != family_filter:
            continue
        if category_filter and clean_text(entry.get("category")) != category_filter:
            continue
        if required_fields and not all(entry["row"].get(field) not in (None, "") for field in required_fields):
            continue
        results.append(entry_to_output(entry))
        if len(results) >= limit:
            break
    return results


def dedupe_b_matches(matches: list[dict[str, Any]]) -> list[dict[str, Any]]:
    deduped: list[dict[str, Any]] = []
    seen: set[tuple[Any, ...]] = set()
    for match in matches:
        key = (
            match.get("material"),
            match.get("family"),
            tuple(match.get("function_tags") or []),
            match.get("function_raw"),
            match.get("row_number"),
        )
        if key not in seen:
            seen.add(key)
            deduped.append(match)
    return deduped


def lookup_b(store: dict[str, Any], materials: list[str]) -> tuple[list[dict[str, Any]], list[str]]:
    results: list[dict[str, Any]] = []
    warnings: list[str] = []
    for material in materials:
        token = normalize_material_token(material)
        matches = dedupe_b_matches([store["entries"][idx] for idx in store["index"].get(token, [])])
        if not matches:
            warnings.append(f"B剂分类表未匹配到材料: {material}")
        results.append(
            {
                "source": "b_classification",
                "requested_material": material,
                "normalized_requested_material": token,
                "matched": bool(matches),
                "match_count": len(matches),
                "matches": matches,
            }
        )
    return results, warnings


def resolve_lookup_result(result: dict[str, Any]) -> dict[str, Any]:
    matches = result.get("matches") or []
    if not matches:
        return {"status": "unmatched", "family": None, "function_tags": [], "function_raw": None, "message": "分类表未命中，无法自动判定。"}
    unique: dict[tuple[str | None, tuple[str, ...], str | None], dict[str, Any]] = {}
    for match in matches:
        key = (match.get("family"), tuple(match.get("function_tags") or []), match.get("function_raw"))
        unique.setdefault(key, match)
    if len(unique) > 1:
        return {"status": "ambiguous", "family": None, "function_tags": [], "function_raw": None, "message": "分类表存在多种匹配，需人工确认。"}
    match = next(iter(unique.values()))
    return {
        "status": "classified",
        "family": match.get("family"),
        "function_tags": match.get("function_tags") or [],
        "function_raw": match.get("function_raw"),
        "message": "分类表命中。",
    }


def exception_rule_for(source_family: str | None, target_family: str | None, function_tags: list[str]) -> dict[str, Any] | None:
    if not source_family or not target_family or source_family == target_family:
        return None
    function_tag_set = set(function_tags or [])
    for rule in CROSS_SYSTEM_EXCEPTION_RULES:
        if rule["source_family"] != source_family:
            continue
        if target_family not in rule["allowed_target_families"]:
            continue
        if function_tag_set.intersection(rule["required_function_tags"]):
            return rule
    return None


def validate_b(store: dict[str, Any], materials: list[str], target_family: str | None) -> tuple[dict[str, Any], list[str]]:
    lookup_results, warnings = lookup_b(store, materials)
    normalized_target = normalize_family_label(target_family) if target_family else None
    assessments = []
    statuses = []
    for result in lookup_results:
        resolved = resolve_lookup_result(result)
        rule = exception_rule_for(resolved.get("family"), normalized_target, resolved.get("function_tags") or [])
        if resolved["status"] != "classified":
            status = resolved["status"]
            reason = resolved["message"]
        elif normalized_target is None or resolved.get("family") == normalized_target:
            status = "compatible"
            reason = "目标体系为空或与分类体系一致。"
        elif rule:
            status = "conditional"
            reason = rule["message"]
        else:
            status = "blocked"
            reason = "不在跨体系例外规则内，不能自动判定为兼容。"
        statuses.append(status)
        assessments.append(
            {
                "requested_material": result["requested_material"],
                "classification": resolved,
                "target_family": normalized_target,
                "status": status,
                "exception_rule": rule,
                "reason": reason,
                "matches": result.get("matches") or [],
            }
        )
    if not assessments:
        overall = "no_materials"
    elif any(status in {"blocked", "ambiguous", "unmatched"} for status in statuses):
        overall = "blocked"
    elif any(status == "conditional" for status in statuses):
        overall = "conditional"
    else:
        overall = "compatible"
    return {
        "source": "b_classification",
        "overall_status": overall,
        "target_family": normalized_target,
        "materials": assessments,
    }, warnings


def apply_status(response: dict[str, Any]) -> dict[str, Any]:
    if response["errors"]:
        response["status"] = "error"
    elif response["warnings"]:
        response["status"] = "warning"
    else:
        response["status"] = "ok"
    return response


def run_tool(payload: dict[str, Any]) -> dict[str, Any]:
    action = clean_text(payload.get("action") or "schema").lower()
    paths = resolve_paths(payload)
    response = base_response(action, paths)
    response["query_summary"].update(
        {
            "materials": as_string_list(payload.get("materials")),
            "cas_numbers": as_string_list(payload.get("cas_numbers")),
            "query": clean_text(payload.get("query")),
            "field": clean_text(payload.get("field") or "all"),
            "family": clean_text(payload.get("family")),
            "category": clean_text(payload.get("category")),
            "has_fields": as_string_list(payload.get("has_fields")),
            "target_family": clean_text(payload.get("target_family")),
            "limit": limit_from(payload.get("limit"), 10),
            "include_all_fields": as_bool(payload.get("include_all_fields")),
        }
    )

    try:
        if action == "schema":
            response["profile"]["supported_actions"] = ["schema", "lookup", "search", "filter", "validate"]
            response["results"].append(master_schema(load_master_store(paths["master_map"])))
            response["safety_results"].append(b_schema(load_b_store(paths["b_classification"])))
        elif action == "lookup":
            materials = response["query_summary"]["materials"]
            cas_numbers = response["query_summary"]["cas_numbers"]
            if not materials and not cas_numbers:
                response["warnings"].append("lookup requires at least one material or CAS number.")
            master_results, master_warnings = lookup_master(
                load_master_store(paths["master_map"]),
                materials,
                cas_numbers,
                response["query_summary"]["include_all_fields"],
            )
            b_results, b_warnings = lookup_b(load_b_store(paths["b_classification"]), materials)
            response["results"] = master_results
            response["safety_results"] = b_results
            response["warnings"].extend(master_warnings + b_warnings)
        elif action == "search":
            query = response["query_summary"]["query"]
            field = response["query_summary"]["field"]
            if field not in {"all", "material", "chemical", "cas", "text"}:
                field = "all"
                response["warnings"].append("Unsupported field; defaulted to all.")
            if not query:
                response["warnings"].append("search requires a non-empty query.")
            else:
                response["results"] = search_master(load_master_store(paths["master_map"]), query, field, response["query_summary"]["limit"])
        elif action == "filter":
            response["results"] = filter_master(
                load_master_store(paths["master_map"]),
                response["query_summary"]["family"],
                response["query_summary"]["category"],
                response["query_summary"]["has_fields"],
                response["query_summary"]["limit"],
            )
        elif action == "validate":
            materials = response["query_summary"]["materials"]
            if not materials:
                response["warnings"].append("validate requires at least one material.")
            validation, validation_warnings = validate_b(
                load_b_store(paths["b_classification"]),
                materials,
                response["query_summary"]["target_family"],
            )
            response["safety_results"] = [validation]
            response["warnings"].extend(validation_warnings)
        else:
            response["errors"].append(f"Unsupported action: {action}")
    except Exception as exc:
        response["errors"].append(f"{type(exc).__name__}: {exc}")

    return apply_status(response)
